import sys
import os
import codecs
import xlrd
from openpyxl import load_workbook
from os import listdir

def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0,len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active()
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(dst_file_path)

def convertXlsxToRst(infn,f, startFromRow=0,endRow = 2000,ncol =2,custom_headers = []):
    wb = load_workbook(infn)

    # grab the active worksheet
    ws = wb.active
    rows = []
    for i,row in enumerate(ws):

        cells = []
        for c in row:
            if c.value is None:
                cells.append("")
                continue
            description = c.value
            if type(description) == str:
                    value = unicode(description,"utf-8")
            else:
                value = unicode(description)

            cells.append(value)
        rows.append(cells)


    if ( not custom_headers ):
        headers = rows[startFromRow]
    else:
        headers = custom_headers
    print "headers", headers
    firstHeader = 0
    for i,r in enumerate(headers):
        if(len(r)>0 ):
            break
        firstHeader = firstHeader+1

    size  = [200]* ncol
    maxCol = 400


    maxHeader = max([(len(u''+x),x) for x in headers])

    for row in rows:
        for i in range((ncol-1),len(headers)-1):

            lenRow_i = len(u''+row[i])

            maxCol=max(maxCol, (lenRow_i+maxHeader[0]+20))



    print "maxCol:", maxCol
    size[ncol-1] = maxCol
    row_separator=('+'+'{:}+'*ncol).format(*[x*"-" for x in size]);

    fmt_row = u'|'+ '|'.join(["{:"+str(x)+"}" for x in size])+"|"

    print >>f, row_separator
    v= headers[:ncol]

    print >>f, fmt_row.format(*v)
    print >>f, ('+'+'{:}+'*ncol).format(*[x*"=" for x in size]);

    firstDataRow  = startFromRow+1

    selected_rows = rows[firstDataRow:endRow]
    if(endRow < 2000):
        print "row",selected_rows


    for row in selected_rows:
        if(endRow < 2000):
            print "row_to_print", row


        '''if not row[firstDataRow]:
            print "not first data"
            continue
        '''
        for i,v in enumerate(row):
                row[i] = v.replace("\n", " ")


        #print "row ", i, row, len(row)
        print >>f,fmt_row.format(*row[firstHeader:(firstHeader+ncol)])

        for i, other_row in enumerate(row[firstHeader+ncol::]):
            if i ==0 :
                print >>f, fmt_row.format(*(['']*ncol))

            if len(other_row)>0:
                key = ""+headers[firstHeader+ncol+i]+ ": " if headers[firstHeader+ncol+i]   else  ""
                v = [""]*(ncol-1) +["  - "+key+  other_row]
                print >>f, fmt_row.format(*v)

        print >>f, row_separator

def getXlsxFiles(path_to_dir, suffix=".xlsx" ):
    filenames = listdir(path_to_dir)
    return [ filename for filename in filenames if filename.endswith( suffix ) ]

if __name__ == "__main__":
    if len(sys.argv)<2 :
        print "Usage " + __file__ + " [Directory for xlsx files] [Directory for xlsx output]"
        quit()
    xlxs_folder =sys.argv[1]
    xlsx= getXlsxFiles(xlxs_folder)

    for xls in xlsx:
        file_name = (xls[0:-5]).lower()

        section=file_name[:(file_name.index("_"))]
        title =  (file_name[(file_name.index("_")+1)::]).replace("_"," ").title()

        f = codecs.open(sys.argv[2]+file_name+".rst", "w", "utf-8")
        convertXlsxToRst(xlxs_folder+xls, f)
        f.close()
